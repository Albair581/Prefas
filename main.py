import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from openpyxl.styles import Border, Side
import pandas as pd
from datetime import datetime, timedelta
import mailing
import calendar

# DOCUMENTATION
# 
# loading workbook
# wb = load_workbook("FILE_NAME.xlsx")
# 
# Save the file
# wb.save("FILE_NAME.xlsx")
# 
# loading worksheets
# active one:
# ws = wb.active
# specific one:
# ws = wb["SHEET_NAME"]
# 
# changing one cell value
# * "cell_num" e.x. "A1", "B2"
# ws["cell_num"] = "value"
# 
# creating sheets
# wb.create_sheet("new_sheet_name")
# 
# change worksheet name
# ws.title = "name"
# 
# append a row of data after the last data
# ws.append(["data1", "data2", "data3", "data4"])
# 
# access or edit a range of cell's data
# for row in range("start row num in integer", "end row num + 1 in integer"):
#   for column in range("start col num in integer", "end col num + 1 in integer"):
#       col = get_column_letter(column)
#       do something to the cell ws[col + str(row)]
# 
# merge cells
# * range e.x. "A1:D3" --> A1 to D3 (a rect)
# ws.merge_cells("range")
# 
# unmerge cells
# ws.unmerge_cells("range")
# 
# insert row
# * replace row with the row that you want to be empty
# ws.insert_rows(row)
# 
# insert column
# * replace col with the col that you want to be empty e.x. 1, 2 (cannot be string like "A", "B")
# ws.insert_cols(col)
# 
# delete row and/or column
# row:
# ws.delete_rows(row_to_be_deleted)
# column:
# ws.delete_cols(column_to_be_deleted)
# 
# move a range of cells
# ws.move_range("range", rows="how many rows to move in integer", cols="same as rows param")
# 
# formulas
# ws["cell"] = "=FORMULA_NAME(parameters)"
# 
# changing cell styles
# cell.font = Font(name="font-family", size="size in integer", 
#             strike=True, color="aRGB color", bold=True,
#             italic=True)
# 
# cell.fill = PatternFill(fill_type=None,
#                 start_color='aRGB color',
#                 end_color='aRGB color')
# cell.border = Border(left=Side(border_style=None,
#                           color='aRGB color'),
#                 right=Side(border_style=None,
#                            color='aRGB color'),
#                 top=Side(border_style=None,
#                          color='aRGB color'),
#                 bottom=Side(border_style=None,
#                             color='aRGB color'),
#                 diagonal=Side(border_style=None,
#                               color='aRGB color'),
#                 diagonal_direction=0,
#                 outline=Side(border_style=None,
#                              color='aRGB color'),
#                 vertical=Side(border_style=None,
#                               color='aRGB color'),
#                 horizontal=Side(border_style=None,
#                                color='aRGB color')
#                )
# cell.alignment = Alignment(horizontal='general',
#                         vertical='bottom',
#                         text_rotation=0,
#                         wrap_text=False,
#                         shrink_to_fit=False,
#                         indent=0)


# TO-DO
# 完善Gmail通知Maggie功能
#
# 境外津貼申請 -> 天數 
# --> 月份最大的一天 - 請假天數 = 津貼天數
# --> 與請假功能連動 (請假時隨時更新)
#
# 境外請假
# --> 8小時算一天
# --> 小時
# --> 津貼是轉換
#
# 境外Belinda
# --> 選擇是否有出差
# --> 選擇出差天數
# --> 計算出差天數
#
# 特休 --> 小時
# 境外 --> 天數


st.set_page_config(
    page_title="新明請假",
    page_icon="💻",
    layout="wide",
    initial_sidebar_state="auto"
)

if "page" not in st.session_state:
    st.session_state.page = 0
if "luser" not in st.session_state:
    st.session_state.luser = "user"
if "lclass" not in st.session_state:
    st.session_state.lclass = "class"
if "atype" not in st.session_state:
    st.session_state.atype = "user"
if "authen" not in st.session_state:
    st.session_state.authen = False
if "abtt" not in st.session_state:
    st.session_state.abtt = ""

st.html("""
            <style>
                [alt=Logo] {
                height:7.5rem;
                }
            </style>
                    """)
st.logo("assets/logo.png")


# authenticated = False
# auth_type = "user"
# auth_user = "defaultU"
# auth_class = "defaultC"

info_board = st.empty()
auth_page = st.empty()
authOut_page = st.empty()
secur = st.empty()

def auth(luser, lclass, outp):
    global auth_page, info_board
    if st.secrets["data"]["userdata"][luser] == lclass:
        st.session_state.luser = luser
        st.session_state.lclass = lclass
        outp.empty()
        auth_page.empty()
        info_board.empty()
        st.session_state.authen = True
        if luser == st.secrets["managers"]["absence"]:
            st.session_state.atype = "absence"
        elif luser == st.secrets["managers"]["design"]:
            st.session_state.atype = "design"
        elif luser == st.secrets["managers"]["other"]:
            st.session_state.atype = "other"
        elif luser == st.secrets["managers"]["ceo"]:
            st.session_state.atype = "ceo"
        elif luser == st.secrets["managers"]["test"]:
            st.session_state.atype = "test"
        else:
            st.session_state.atype = "user"
        st.session_state.page = 2
        st.rerun()
    else:
        st.error("登入資訊錯誤! 請選擇正確使用者及部門!", icon="❌")

def append_data(user:str, wclass:str, type: str, sdate: str, stime: str, edate: str, etime: str, tt: str, approve_s: str, ap_who: str, special: str, out: str, loadTimes: int, outmoney):
    wb = load_workbook("userData.xlsx")
    ws = wb["請假單"]
    ws.append([user, wclass, type, sdate, stime, edate, etime, tt, approve_s, ap_who, special, out, loadTimes, outmoney])
    wb.save("userData.xlsx")

def remove_all_data():
    wb = load_workbook("userData.xlsx")
    ws = wb["請假單"]
    for row in range(2, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            col = get_column_letter(column)
            ws[col + str(row)].value = None
    wb.save("userData.xlsx")
    message_text = f"""
!警告!----
{st.session_state.luser} 已在 
{datetime.now().replace(microsecond=0)}
刪除了所有請假資料!!!!

Prefas請假網頁警告系統
"""
    if st.session_state.atype == "absence":
        mailing.send_email("刪除資料警告", "alvinhuang0709@gmail.com", message_text)
    elif st.session_state.atype == "test":
        mailing.send_email("刪除資料警告", "albertoyucheng@gmail.com", message_text)

def log_out():
    st.session_state.luser = "default"
    st.session_state.lclass = "default"
    st.session_state.authen = False
    st.session_state.atype = "user"
    st.session_state.page = 1

def enter(key:str):
    global secur
    if st.secrets["webSecurity"]["key"] == key:
        if st.secrets["webSecurity"]["pass_way"] == "KEY" and st.secrets["webSecurity"]["pass_type"] == "DOOR":
            secur.empty()
            st.session_state.page = 1
        else:
            st.error("系統錯誤105, 請通知網站維護人員!", icon="❌")
    else:
        st.error("芝麻沒開門! 請輸入正確密碼!", icon="❌")

def authOut(luser:str):
    global authOut_page, auth_page, info_board
    st.session_state.luser = luser
    if not st.session_state.luser == "董事長Andy":
        st.session_state.lclass = "境外"
    else:
        st.session_state.lclass = "董事"
    authOut_page.empty()
    auth_page.empty()
    info_board.empty()
    st.session_state.atype = "ceo" if st.session_state.luser == "董事長Andy" else "user"
    st.session_state.authen = True
    st.session_state.page = 2
    st.rerun()

def calculate_working_hours(start_datetime_str, end_datetime_str):
    # Define workday start, end, and lunch break times
    workday_start_time = "08:30:00"
    workday_end_time = "17:30:00"
    lunch_start_time = "12:00:00"
    lunch_end_time = "13:00:00"
    
    workday_start = datetime.strptime(workday_start_time, "%H:%M:%S").time()
    workday_end = datetime.strptime(workday_end_time, "%H:%M:%S").time()
    lunch_start = datetime.strptime(lunch_start_time, "%H:%M:%S").time()
    lunch_end = datetime.strptime(lunch_end_time, "%H:%M:%S").time()

    # Parse input datetime strings
    start_datetime = datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S")
    end_datetime = datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M:%S")

    if start_datetime >= end_datetime:
        return 0

    # Initialize total working hours
    total_working_hours = 0.0

    current_datetime = start_datetime

    while current_datetime.date() <= end_datetime.date():
        day_start = datetime.combine(current_datetime.date(), workday_start)
        day_end = datetime.combine(current_datetime.date(), workday_end)
        lunch_start_dt = datetime.combine(current_datetime.date(), lunch_start)
        lunch_end_dt = datetime.combine(current_datetime.date(), lunch_end)

        if current_datetime.date() == start_datetime.date():
            if current_datetime.time() > workday_start:
                day_start = current_datetime
            if start_datetime.time() > lunch_start and start_datetime.time() < lunch_end:
                lunch_start_dt = start_datetime
        if current_datetime.date() == end_datetime.date():
            day_end = end_datetime
            if end_datetime.time() > lunch_start and end_datetime.time() < lunch_end:
                lunch_end_dt = end_datetime

        # Calculate working hours for the current day
        day_working_seconds = (day_end - day_start).total_seconds()
        lunch_seconds = (lunch_end_dt - lunch_start_dt).total_seconds()

        if day_start < lunch_start_dt and day_end > lunch_end_dt:
            day_working_seconds -= lunch_seconds
        elif day_start < lunch_start_dt and day_end > lunch_start_dt:
            day_working_seconds -= (day_end - lunch_start_dt).total_seconds()
        elif day_start < lunch_end_dt and day_end > lunch_end_dt:
            day_working_seconds -= (lunch_end_dt - day_start).total_seconds()

        total_working_hours += day_working_seconds / 3600

        current_datetime = day_start + timedelta(days=1)
        current_datetime = datetime.combine(current_datetime.date(), workday_start)

    return total_working_hours

def convert_to_days(hours, hours_per_day=8):
    total_days = round((hours / hours_per_day), 2)
    return total_days

def remove_all_outmoney():
    wb = load_workbook("userData.xlsx")
    ws = wb["境外津貼"]
    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"] = "?"
        ws[f"C{row}"] = 0
    wb.save("userData.xlsx")

def remove_all_overseas():
    wb = load_workbook("userData.xlsx")
    ws = wb["出差資料"]
    for row in range(2, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            col = get_column_letter(column)
            ws[f"{col}{row}"].value = None
    wb.save("userData.xlsx")

@st.experimental_dialog("刪除資料密碼確認")
def ask_remove_outmoney(type):
    types = ["outmoney", "absencedata", "overseas"]
    assert type in types
    remove_pass = st.text_input("刪除密碼", type="password")
    if st.button("確認"):
        if remove_pass == st.secrets["webSecurity"]["remove_outmoney_key"]:
            if type == "outmoney":
                remove_all_outmoney()
                st.write("已刪除資料, 可關掉此頁面!")
            elif type == "absencedata":
                remove_all_data()
                st.write("已刪除資料, 可關掉此頁面!")
            elif type == "overseas":
                remove_all_overseas()
                st.write("已刪除資料, 可關掉此頁面!")
        else:
            st.error("密碼錯誤", icon="🚨")

def calculate_days_between(start_date_str, end_date_str, date_format='%Y-%m-%d'):
    start_date = datetime.strptime(start_date_str, date_format)
    end_date = datetime.strptime(end_date_str, date_format)
    date_difference = end_date - start_date
    return date_difference.days

# password website security
if st.session_state.page == 0:
    with secur.container():
        col1, col2= st.columns(2)

        col1.image("assets/dress.png")
        with col2:
            st.html("<h1 style='font-size: 55px;'>Welcome, Pretty.</h1>")
            passkey = st.text_input("輸入密碼", max_chars=int(st.secrets["webSecurity"]["max_log"]), type="password")
            if st.button("芝麻開門!"):
                enter(str(passkey))
    st.html(f"<h4 style='text-align: center;'>© 2024 Copyright Albert Huang 😊 版本Prefas {st.secrets["cloud_version"]}</h4>")
        # st.write("©2024 Copyright Albert Huang :smile:")
            
# login
if st.session_state.page == 1 and not st.session_state.authen:
    with info_board.container():
        st.header("Prefas請假網站公告")
        st.write(":rainbow[請注意!] 檢核資料時:orange[會]發送郵件到人事部, 請:red[**勿**]登入他人帳號--違者:red[**必究**]!")

    with auth_page.container():
        st.title("新名請假系統登入")
        log_user = st.selectbox("請選您的名字", list(st.secrets["user"]["userlist"]))
        log_class = st.selectbox("請選您的部門", list(st.secrets["user"]["userclass"]))
        log_btn = st.button("登入", key="in")
        if log_btn:
            auth(str(log_user), str(log_class), authOut_page)
    with authOut_page.container():
        st.title("境外請假系統登入")
        log_user = st.selectbox("請選您的名字", list(st.secrets["user"]["outlist"]))
        log_btn = st.button("登入", key="out")
        if log_btn:
            authOut(str(log_user))
if st.session_state.page == 2:
    with st.sidebar:
        st.header("登入資訊", divider="rainbow")
        st.subheader(f"使用者: {st.session_state.luser}")
        class_text = f"部門: {st.session_state.lclass}部" if st.session_state.atype != "ceo" else f"{st.session_state.lclass}長"
        st.subheader(class_text)
        st.subheader(f"登入狀態: {"已登入" if st.session_state.authen else "未登入"}")
        st.write("請按兩次登出!")
        if st.button("登出"):
            log_out()
    st.title(f"新名請假系統 - {st.session_state.luser}")
    if st.session_state.atype != "absence" and st.session_state.atype != "ceo" and st.session_state.atype != "test":
        if st.session_state.lclass != "境外" and not st.session_state.atype == "other":
            abtype = st.selectbox("請假類別", 
                                ["特休", "病假", "事假", "喪假", "婚假", "生育假"])
        else:
            abtype = st.selectbox("請假類別", 
                                ["特休", "境外", "病假", "事假", "喪假", "婚假", "生育假"])

        absdate = st.date_input("請假開始日期")
        abstime = st.time_input("請假開始小時 (24時制)")
        abstime = str(abstime)

        abedate = st.date_input("請假結束日期")
        abetime = st.time_input("請假結束小時 (24時制)")
        abetime = str(abetime)

        start_string = f"{str(absdate)} {str(abstime)}"
        end_string = f"{str(abedate)} {str(abetime)}"
        
        tt = calculate_working_hours(start_string, end_string)
        st.session_state.abtt = str(round(tt, 2))
        abtt = st.text_input("總共請假小時", key="abtt")

        ok_btn = st.button("送出檢核")
        if ok_btn:
            wb = load_workbook("userData.xlsx")
            ws = wb["境外津貼"]
            if st.session_state.lclass == "境外" or st.session_state.atype == "other":
                for row in range(2, ws.max_row):
                    cell = f"C{row}"
                    if ws[cell].value == 0:
                        ws[cell] = 1
                        dtoday = datetime.today()
                        year = dtoday.year
                        month = dtoday.month
                        _, max_day = calendar.monthrange(year, month)
                        ws[f"B{row}"] = int(max_day)
                outmoney = ((int(ws[f"B{st.secrets["data"]["outmoneyList"][st.session_state.luser]}"].value) * 8) - int(tt))
                outmoney = convert_to_days(outmoney)
                ws[f"B{st.secrets["data"]["outmoneyList"][st.session_state.luser]}"] = outmoney
            else:
                outmoney = "無"
            wb.save("userData.xlsx")
            # for row in range(1, ws.max_row + 1):
            #     if ws[f"A{row}"].value == str(st.session_state.luser):
            #         outmoney = ws[f"B{row}"].value
            append_data(st.session_state.luser, st.session_state.lclass, abtype, str(absdate), str(abstime), str(abedate), str(abetime), float(abtt), "未檢核", "未檢核", "無特休假", "無境外假", 0, outmoney)
            st.success("成功送出請假資料!", icon="✅")

    if st.session_state.atype == "other":
        st.header("境外津貼選項")
        sgoing = st.date_input("出差開始日期")
        egoing = st.date_input("出差結束日期")
        td = calculate_days_between(str(sgoing), str(egoing))
        dates = st.number_input("總出差天數", value=int(td))
        if st.button("確認資料"):
            wb = load_workbook("userData.xlsx")
            ws = wb["境外津貼"]
            if ws[f"A{st.secrets["data"]["outmoneyList"][str(st.session_state.luser)]}"].value == str(st.session_state.luser):
                ws[f"B{st.secrets["data"]["outmoneyList"][str(st.session_state.luser)]}"] = int(dates)
                ws[f"C{st.secrets["data"]["outmoneyList"][str(st.session_state.luser)]}"] = 1
            ws = wb["出差資料"]
            ws.append([st.session_state.luser, str(sgoing), str(egoing), int(dates), "未檢核", "未檢核", 0])
            wb.save("userData.xlsx")

    if st.session_state.atype == "absence" or st.session_state.atype == "test":
        if st.button("清除資料", type="primary"):
            ask_remove_outmoney(type="absencedata")

        with open("userData.xlsx", "rb") as template_file:
            template_byte = template_file.read()

        st.download_button(label="下載請假資料(格式: Excel)",
                            data=template_byte,
                            file_name="userData.xlsx",
                            mime='application/octet-stream')
        
        if st.button("清除境外津貼資料", type="primary"):
            ask_remove_outmoney(type="outmoney")

        if st.button("清除出差資料", type="primary"):
            ask_remove_outmoney(type="overseas")
    

    st.header("檢核報告")
    if st.session_state.atype != "user" and st.session_state.atype != "test":
        lb = load_workbook("userData.xlsx")
        ls = lb["請假單"]
        lr = ls.max_row

        actual_user = st.session_state.atype

        dataList = []

        for row in range(2, ls.max_row + 1):
            if actual_user == "other":
                if ls[f"A{row}"].value != "總經理Belinda" and ls[f"B{row}"].value != "境外":
                    dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                     "status":ls[f"I{row}"].value,
                                    "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                    "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                    "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
                elif str(ls[f"J{row}"].value) == "未檢核" and ls[f"A{row}"].value == "總經理Belinda":
                    dataList.append({"approved":False,
                                     "status":None,
                                    "user":"自己", "class":None, "type":None, 
                                    "sdate":None, "stime":None, "edate":None,
                                    "etime":None, "ttime":None, "ap_who":None,
                                    "special":None, "out": None})
                elif ls[f"B{row}"].value == "境外":
                    dataList.append({"approved":False,
                                     "status":None,
                                    "user":"境外部", "class":None, "type":None, 
                                    "sdate":None, "stime":None, "edate":None,
                                    "etime":None, "ttime":None, "ap_who":None,
                                    "special":None, "out": None})
                elif str(ls[f"J{row}"].value) == "董事長Andy" and ls[f"A{row}"].value == "總經理Belinda":
                    dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                     "status":ls[f"I{row}"].value,
                                    "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                    "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                    "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
                else:
                    pass
            elif actual_user == "design":
                if ls[f"A{row}"].value != "溫金蘭" and ls[f"B{row}"].value == "設計":
                    dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                     "status":ls[f"I{row}"].value,
                                    "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                    "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                    "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
                elif str(ls[f"J{row}"].value) == "未檢核" and ls[f"A{row}"].value == "溫金蘭":
                    dataList.append({"approved":False,
                                     "status":None,
                                    "user":"自己", "class":None, "type":None, 
                                    "sdate":None, "stime":None, "edate":None,
                                    "etime":None, "ttime":None, "ap_who":None,
                                    "special":None, "out": None})
                elif ls[f"B{row}"].value != "設計":
                    dataList.append({"approved":False,
                                     "status":None,
                                    "user": "非設計部", "class":None, "type":None, 
                                    "sdate":None, "stime":None, "edate":None,
                                    "etime":None, "ttime":None, "ap_who":None,
                                    "special":None, "out": None})
                else:
                    dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                     "status":ls[f"I{row}"].value,
                                    "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                    "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                    "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
            elif actual_user == "absence":
                dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                 "status":ls[f"I{row}"].value,
                                "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
            elif actual_user == "ceo":
                if ls[f"B{row}"].value == "境外" or ls[f"A{row}"].value == "總經理Belinda":
                    dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                     "status":ls[f"I{row}"].value,
                                    "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                    "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                    "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                    "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})
                elif ls[f"B{row}"].value != "境外":
                    dataList.append({"approved":False,
                                     "status":None,
                                    "user": "非境外部", "class":None, "type":None, 
                                    "sdate":None, "stime":None, "edate":None,
                                    "etime":None, "ttime":None, "ap_who":None,
                                    "special":None, "out":None})

        df = pd.DataFrame(dataList)
        st.caption(":blue[打勾時, 勾與勾之間需要間隔1秒, 讓系統更新!]")
        st.caption(":red[一旦檢核後, 無法更改! (請注意)]")

        final_df = st.data_editor(
            df,
            column_config={
                "approved": "檢核狀態",
                "status": "檢核狀態",
                "user": "請假員工",
                "class": "部門",
                "type": "請假類別",
                "sdate": "開始日期",
                "stime": "開始時間",
                "edate": "結束日期",
                "etime": "結束時間",
                "ttime": "總請假時間",
                "ap_who": "檢核人員",
                "special": "剩餘特休假",
                "out": "剩餘境外假",
            },
            disabled=["status", "user", "class", "type", "sdate", "stime", "edate", "etime", "ttime", "ap_who", "special", "out"],
            hide_index=True,
        )
        # approved_len = len(list(final_df.loc[final_df["approved"]]["approved"]))
        approve_list = []
        for row in range(2, ls.max_row + 1):
            approved = final_df["approved"][row - 2]
            ap_bool = "True" if approved else "False"
            status = "已檢核" if ap_bool == "True" else "未檢核"
            approve_list.append(ap_bool)
            if not dataList[row - 2]["class"] == None:
                if status == "已檢核": 
                    ls[f"I{row}"] = status
                    ls[f"M{row}"] = int(ls[f"M{row}"].value) + 1
                    if st.session_state.atype == "ceo" and ls[f"M{row}"].value == 1:
                        message_to_send = f"""
Dear Maggie 萬菁,
    此郵件已發到您的郵箱是因為:
董事長Andy在 {str(datetime.now().replace(microsecond=0))} 已檢核了
{str(ls[f"A{row}"].value)} 在Excel表中 第{str(row)}行 的請假。
請確認。

From,
Prefas請假網站系統
                        """
                        mailing.send_email(subject="董事長檢核報告", recipient=st.secrets["mailing"]["maggie"], message_text=message_to_send)
                if status == "已檢核" and ls[f"J{row}"].value == "未檢核" and st.session_state.atype != "user":
                    ls[f"J{row}"] = str(st.session_state.luser)

        # print(approve_list)

        lb.save("userData.xlsx")
    elif st.session_state.atype == "user":
        lb = load_workbook("userData.xlsx", read_only=True)
        ls = lb["請假單"]
        # lr = ls.max_row

        dataList = []

        for row in range(2, ls.max_row + 1):
            if ls[f"A{row}"].value == st.session_state.luser:
                dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})

        df = pd.DataFrame(dataList)
        st.caption("檢核資料可能要重新載入後才會生效!")
        final_df = st.data_editor(
            df,
            column_config={
                "approved": "檢核狀態",
                "user": "請假員工",
                "class": "部門",
                "type": "請假類別",
                "sdate": "開始日期",
                "stime": "開始時間",
                "edate": "結束日期",
                "etime": "結束時間",
                "ttime": "總請假時間",
                "ap_who": "檢核人員",
                "special": "剩餘特休假",
                "out": "剩餘境外假",
            },
            disabled=["approved", "status", "user", "class", "type", "sdate", "stime", "edate", "etime", "ttime", "ap_who", "special", "out"],
            hide_index=True,
        )

        lb.close()

    elif st.session_state.atype == "test":
        lb = load_workbook("userData.xlsx", read_only=True)
        ls = lb["請假單"]

        dataList = []

        for row in range(2, ls.max_row + 1):
            if ls[f"A{row}"].value == st.session_state.luser:
                dataList.append({"approved":False if ls[f"I{row}"].value != "已檢核" else True,
                                "user":ls[f"A{row}"].value, "class":ls[f"B{row}"].value, "type":ls[f"C{row}"].value, 
                                "sdate":ls[f"D{row}"].value, "stime": ls[f"E{row}"].value, "edate":ls[f"F{row}"].value,
                                "etime":ls[f"G{row}"].value, "ttime":ls[f"H{row}"].value, "ap_who":str(ls[f"J{row}"].value),
                                "special":ls[f"K{row}"].value, "out": ls[f"L{row}"].value})

        df = pd.DataFrame(dataList)
        st.caption("檢核資料可能要重新載入後才會生效!")
        final_df = st.data_editor(
            df,
            column_config={
                "approved": "檢核狀態",
                "user": "請假員工",
                "class": "部門",
                "type": "請假類別",
                "sdate": "開始日期",
                "stime": "開始時間",
                "edate": "結束日期",
                "etime": "結束時間",
                "ttime": "總請假時間",
                "ap_who": "檢核人員",
                "special": "剩餘特休假",
                "out": "剩餘境外假",
            },
            disabled=["approved", "status", "user", "class", "type", "sdate", "stime", "edate", "etime", "ttime", "ap_who", "special", "out"],
            hide_index=False,
        )

        lb.close()

    if st.session_state.atype == "ceo":
        st.header("出差津貼申請")
        lb = load_workbook("userData.xlsx")
        ls = lb["出差資料"]
        dataList = []
        for row in range(2, ls.max_row + 1):
            dataList.append({"approved": False if ls[f"E{row}"].value != "已檢核" else True,
                             "user":ls[f"A{row}"].value,
                             "sgoing":ls[f"B{row}"].value,
                             "egoing":ls[f"C{row}"].value,
                             "tdate":ls[f"D{row}"].value,
                             "ap_who":ls[f"F{row}"].value})
        df = pd.DataFrame(dataList)
        st.caption("檢核資料可能要重新載入後才會生效!")
        final_df = st.data_editor(
            df,
            column_config={
                "approved": "檢核狀態",
                "user": "出差員工",
                "sgoing": "開始日期",
                "egoing": "結束日期",
                "tdate": "總出差天數",
                "ap_who": "檢核人員",
            },
            disabled=["user", "sgoing", "egoing", "tdate", "ap_who"],
            hide_index=True,
        )

        approve_list = []
        for row in range(2, ls.max_row + 1):
            approved = final_df["approved"][row - 2]
            ap_bool = "True" if approved else "False"
            status = "已檢核" if ap_bool == "True" else "未檢核"
            approve_list.append(ap_bool)
            if not dataList[row - 2]["user"] == None:
                if status == "已檢核": 
                    ls[f"E{row}"] = status
                    ls[f"G{row}"] = int(ls[f"G{row}"].value) + 1
                    if st.session_state.atype == "ceo" and ls[f"G{row}"].value == 1:
                        message_to_send = f"""
Dear Maggie 萬菁,
    此郵件已發到您的郵箱是因為:
董事長Andy在 {str(datetime.now().replace(microsecond=0))} 已檢核了
{str(ls[f"A{row}"].value)} 在Excel表中 第{str(row)}行 的出差申請。
請確認。

From,
Prefas網站出差申請系統
                        """
                        mailing.send_email(subject="董事長檢核出差報告", recipient=st.secrets["mailing"]["maggie"], message_text=message_to_send)
                if status == "已檢核" and ls[f"E{row}"].value == "已檢核" and st.session_state.atype != "user":
                    ls[f"F{row}"] = str(st.session_state.luser)

        # print(approve_list)

        lb.save("userData.xlsx")

    if st.session_state.atype == "other":
        st.header("出差津貼申請")
        lb = load_workbook("userData.xlsx")
        ls = lb["出差資料"]
        dataList = []
        for row in range(2, ls.max_row + 1):
            dataList.append({"approved": False if ls[f"E{row}"].value != "已檢核" else True,
                             "user":ls[f"A{row}"].value,
                             "sgoing":ls[f"B{row}"].value,
                             "egoing":ls[f"C{row}"].value,
                             "tdate":ls[f"D{row}"].value,
                             "ap_who":ls[f"F{row}"].value})
        df = pd.DataFrame(dataList)
        st.caption("檢核資料可能要重新載入後才會生效!")
        final_df = st.data_editor(
            df,
            column_config={
                "approved": "檢核狀態",
                "user": "出差員工",
                "sgoing": "開始日期",
                "egoing": "結束日期",
                "tdate": "總出差天數",
                "ap_who": "檢核人員",
            },
            disabled=["approved", "user", "sgoing", "egoing", "tdate", "ap_who"],
            hide_index=True,
        )
        lb.close()
